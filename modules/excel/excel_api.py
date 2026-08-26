import io
import json
import re
import os
import tempfile
import zipfile
import shutil
from copy import copy

from flask import Blueprint, jsonify, request, send_file

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - dependency guard
    load_workbook = None

try:
    import win32com.client
    import pythoncom
except ImportError:
    win32com = None

excel_bp = Blueprint('excel_bp', __name__)
CELL_ADDR_RE = re.compile(r"^[A-Z]{1,3}[1-9][0-9]*$")


@excel_bp.route('/apply-updates', methods=['POST'])
def apply_updates():
    """
    Cập nhật giá trị vào file Excel được upload trong khi vẫn giữ nguyên định dạng (styles).
    
    Hỗ trợ hai loại thao tác:
    1. Cập nhật giá trị ô: Dựa trên địa chỉ ô (vd: 'A1').
    2. Chèn dòng mới: Chèn dòng tại vị trí chỉ định và sao chép định dạng từ dòng phía trên.
    
    Returns:
        Response: File Excel đã được cập nhật hoặc lỗi JSON.
    """
    if load_workbook is None:
        return jsonify({"error": "Thiếu thư viện openpyxl trên server."}), 500

    uploaded_file = request.files.get("file")
    if uploaded_file is None:
        return jsonify({"error": "Cần upload file Excel (.xlsx)."}), 400

    filename = uploaded_file.filename or "KetQua_Excel.xlsx"
    if not filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Chỉ hỗ trợ file định dạng .xlsx."}), 400

    updates_raw = request.form.get("updates", "[]")
    try:
        updates = json.loads(updates_raw)
    except json.JSONDecodeError:
        return jsonify({"error": "Dữ liệu updates không phải JSON hợp lệ."}), 400

    if not isinstance(updates, list):
        return jsonify({"error": "Dữ liệu updates phải là mảng JSON."}), 400

    try:
        workbook = load_workbook(filename=io.BytesIO(uploaded_file.read()))
    except Exception:
        return jsonify({"error": "Không đọc được file Excel đầu vào."}), 400

    for idx, item in enumerate(updates):
        if not isinstance(item, dict):
            return jsonify({"error": f"Update tại vị trí {idx} không hợp lệ."}), 400

        sheet_name = str(item.get("sheet", "")).strip()
        if not sheet_name:
            return jsonify({"error": f"Thiếu tên sheet tại vị trí {idx}."}), 400
        if sheet_name not in workbook.sheetnames:
            return jsonify({"error": f"Không tìm thấy sheet '{sheet_name}' trong file Excel."}), 400

        worksheet = workbook[sheet_name]

        if item.get("type") == "insert_row":
            try:
                row_idx = int(item.get("row"))
            except (ValueError, TypeError):
                return jsonify({"error": f"Dòng chèn tại vị trí {idx} không hợp lệ."}), 400
            
            worksheet.insert_rows(row_idx)
            
            # Copy styles and formulas from the row above
            src_row = row_idx - 1
            if src_row > 0:
                # Need to import Translator here or at the top
                from openpyxl.formula.translate import Translator
                for col_idx in range(1, worksheet.max_column + 1):
                    src_cell = worksheet.cell(row=src_row, column=col_idx)
                    tgt_cell = worksheet.cell(row=row_idx, column=col_idx)
                    
                    if src_cell.has_style:
                        tgt_cell.font = copy(src_cell.font)
                        tgt_cell.border = copy(src_cell.border)
                        tgt_cell.fill = copy(src_cell.fill)
                        tgt_cell.number_format = copy(src_cell.number_format)
                        tgt_cell.protection = copy(src_cell.protection)
                        tgt_cell.alignment = copy(src_cell.alignment)
                    
                    if src_cell.data_type == 'f':
                        try:
                            tgt_cell.value = Translator(src_cell.value, origin=src_cell.coordinate).translate_formula(tgt_cell.coordinate)
                        except Exception:
                            # Fallback if formula translation fails
                            tgt_cell.value = src_cell.value
                            
            # Fix all formulas below the inserted row because openpyxl insert_rows doesn't update references
            def shift_formula(formula, insert_row_idx, num_rows=1):
                def repl(m):
                    if m.group(5):  # Range reference
                        col1_abs, col1, row1_abs, row1_str = m.group(1), m.group(2), m.group(3), m.group(4)
                        col2_abs, col2, row2_abs, row2_str = m.group(6), m.group(7), m.group(8), m.group(9)
                        row1 = int(row1_str)
                        row2 = int(row2_str)
                        
                        if row1_abs != '$' and row1 >= insert_row_idx:
                            row1 += num_rows
                            
                        # Expand the range if inserted exactly at boundary (row2 + 1)
                        if row2_abs != '$':
                            if row2 >= insert_row_idx or row2 == insert_row_idx - 1:
                                row2 += num_rows
                                
                        return f"{col1_abs}{col1}{row1_abs}{row1}:{col2_abs}{col2}{row2_abs}{row2}"
                    else:  # Single cell reference
                        col_abs, col, row_abs, row_str = m.group(1), m.group(2), m.group(3), m.group(4)
                        row = int(row_str)
                        if row_abs != '$' and row >= insert_row_idx:
                            row += num_rows
                        return f"{col_abs}{col}{row_abs}{row}"

                pattern = r'(?<![a-zA-Z])(\$?)([A-Z]{1,3})(\$?)(\d+)(?:(:)(\$?)([A-Z]{1,3})(\$?)(\d+))?(?!\()'
                return re.sub(pattern, repl, formula)

            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.row != row_idx and cell.data_type == 'f' and cell.value and isinstance(cell.value, str):
                        new_formula = shift_formula(cell.value, row_idx)
                        if new_formula != cell.value:
                            cell.value = new_formula
                            
            continue

        cell_address = str(item.get("address", "")).strip().upper()
        if not CELL_ADDR_RE.fullmatch(cell_address):
            return jsonify({"error": f"Địa chỉ ô '{cell_address}' không hợp lệ."}), 400

        worksheet[cell_address].value = item.get("value")

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    output_name = (request.form.get("filename", "") or filename).strip() or "KetQua_Excel.xlsx"
    if not output_name.lower().endswith(".xlsx"):
        output_name = f"{output_name}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=output_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@excel_bp.route('/extract-charts', methods=['POST'])
def extract_charts():
    """
    Trích xuất toàn bộ biểu đồ từ file Excel và gom vào thư mục có tên của sheet.
    Trả về file ZIP chứa tất cả các ảnh dưới dạng SVG.
    """
    if win32com is None:
        return jsonify({"error": "Thiếu thư viện pywin32 hoặc không chạy trên môi trường Windows."}), 500

    uploaded_file = request.files.get("file")
    if uploaded_file is None:
        return jsonify({"error": "Cần upload file Excel."}), 400

    filename = uploaded_file.filename
    if not (filename.lower().endswith(".xlsx") or filename.lower().endswith(".xls")):
        return jsonify({"error": "Chỉ hỗ trợ file Excel (.xls, .xlsx)."}), 400

    temp_dir = tempfile.mkdtemp()
    try:
        ext = os.path.splitext(filename)[1]
        input_filepath = os.path.join(temp_dir, f"input_file{ext}")
        uploaded_file.save(input_filepath)
        
        # Khởi tạo COM để dùng pywin32 trong thread của Flask
        pythoncom.CoInitialize()
        
        excel = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            wb = excel.Workbooks.Open(os.path.abspath(input_filepath))
            
            valid_charts_count = 0
            charts_dir = os.path.join(temp_dir, "charts")
            os.makedirs(charts_dir, exist_ok=True)
            
            for sheet in wb.Worksheets:
                chart_objects = sheet.ChartObjects()
                if chart_objects.Count > 0:
                    sheet_name = sheet.Name
                    sheet_dir = os.path.join(charts_dir, sheet_name)
                    os.makedirs(sheet_dir, exist_ok=True)
                    
                    used_names = {}
                    for idx, chart_obj in enumerate(chart_objects):
                        # Bỏ qua các object quá nhỏ (thường là icon, nút bấm, hình rác)
                        if chart_obj.Width < 50 or chart_obj.Height < 50:
                            continue
                            
                        chart = chart_obj.Chart
                        
                        # Tên mặc định của biểu đồ
                        chart_name = f"chart_{idx + 1}"
                        
                        # Thử lấy tiêu đề biểu đồ nếu có
                        try:
                            if chart.HasTitle:
                                title_text = chart.ChartTitle.Text
                                if title_text:
                                    # Loại bỏ các ký tự không hợp lệ trong tên file
                                    clean_title = re.sub(r'[\\/*?:"<>|\r\n\t]', "", title_text).strip()
                                    if clean_title:
                                        chart_name = clean_title[:100]
                        except Exception:
                            pass
                            
                        # Tránh trùng tên file trong cùng sheet
                        base_name = chart_name
                        if base_name in used_names:
                            used_names[base_name] += 1
                            chart_name = f"{base_name}_{used_names[base_name]}"
                        else:
                            used_names[base_name] = 1
                            
                        image_path = os.path.join(sheet_dir, f"{chart_name}.svg")
                        try:
                            chart.Export(os.path.abspath(image_path), "SVG")
                            # Lọc các file SVG lỗi hoặc rỗng (thường < 100 bytes)
                            if os.path.exists(image_path):
                                if os.path.getsize(image_path) < 100:
                                    os.remove(image_path)
                                else:
                                    valid_charts_count += 1
                        except Exception:
                            pass
            
            wb.Close(SaveChanges=False)
            excel.Quit()
        except Exception as e:
            if excel:
                excel.Quit()
            raise e
        finally:
            pythoncom.CoUninitialize()

        if valid_charts_count == 0:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return jsonify({"error": "Không tìm thấy biểu đồ hợp lệ nào (có thể toàn bộ là biểu đồ trống hoặc hình rác)."}), 404

        # Nén thành zip
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
            for root, _, files in os.walk(charts_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, charts_dir)
                    zipf.write(file_path, arcname)

        zip_buffer.seek(0)
        zip_filename = f"Charts_{os.path.splitext(filename)[0]}.zip"
        
        # Cleanup
        shutil.rmtree(temp_dir, ignore_errors=True)
        
        return send_file(
            zip_buffer,
            as_attachment=True,
            download_name=zip_filename,
            mimetype="application/zip"
        )
    except Exception as e:
        shutil.rmtree(temp_dir, ignore_errors=True)
        return jsonify({"error": f"Lỗi khi trích xuất biểu đồ: {str(e)}"}), 500


@excel_bp.route('/export-electricity-report', methods=['POST'])
def export_electricity_report():
    """
    Nhận dữ liệu điện 3 giá dạng JSON và xuất file Excel báo cáo tiêu thụ điện.
    
    Request JSON body:
    {
        "data": [
            {"nam": 2024, "thang": 1, "ky": 1, "thue_vat": 8, "bt_don_gia": 1833, "bt_san_luong": 51559, "cd_don_gia": 3398, "cd_san_luong": 19180, "td_don_gia": 1190, "td_san_luong": 28653, "ghi_chu": ""},
            ...
        ],
        "filename": "BaoCao_Dien_2024.xlsx"
    }
    
    Returns:
        Response: File Excel (.xlsx) đã format đẹp.
    """
    if load_workbook is None:
        return jsonify({"error": "Thiếu thư viện openpyxl trên server."}), 500

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "Thiếu thư viện openpyxl."}), 500

    payload = request.get_json(silent=True)
    if not payload or "data" not in payload:
        return jsonify({"error": "Dữ liệu JSON không hợp lệ. Cần trường 'data'."}), 400

    rows_data = payload["data"]
    if not isinstance(rows_data, list) or len(rows_data) == 0:
        return jsonify({"error": "Mảng 'data' rỗng hoặc không hợp lệ."}), 400

    output_name = payload.get("filename", "BaoCao_TieuThu_Dien.xlsx").strip()
    if not output_name.lower().endswith(".xlsx"):
        output_name += ".xlsx"

    # ── Chuẩn hóa dữ liệu ──
    required_fields = ["nam", "thang", "ky", "bt_don_gia", "bt_san_luong",
                       "cd_don_gia", "cd_san_luong", "td_don_gia", "td_san_luong"]
    cleaned = []
    for idx, row in enumerate(rows_data):
        if not isinstance(row, dict):
            return jsonify({"error": f"Dòng {idx} không phải object."}), 400
        entry = {}
        for f in required_fields:
            val = row.get(f)
            try:
                entry[f] = float(val) if val is not None else 0
            except (ValueError, TypeError):
                entry[f] = 0
        try:
            entry["thue_vat"] = float(row.get("thue_vat", 0)) if row.get("thue_vat") is not None else 0
        except (ValueError, TypeError):
            entry["thue_vat"] = 0

        entry["ghi_chu"] = str(row.get("ghi_chu", ""))
        
        # Tính toán
        entry["bt_thanh_tien"] = entry["bt_don_gia"] * entry["bt_san_luong"]
        entry["cd_thanh_tien"] = entry["cd_don_gia"] * entry["cd_san_luong"]
        entry["td_thanh_tien"] = entry["td_don_gia"] * entry["td_san_luong"]
        entry["tong_san_luong"] = entry["bt_san_luong"] + entry["cd_san_luong"] + entry["td_san_luong"]
        entry["tien_chua_thue"] = entry["bt_thanh_tien"] + entry["cd_thanh_tien"] + entry["td_thanh_tien"]
        entry["tien_thue"] = round(entry["tien_chua_thue"] * (entry["thue_vat"] / 100))
        entry["tong_thanh_tien"] = entry["tien_chua_thue"] + entry["tien_thue"]
        cleaned.append(entry)

    # ── Tổng hợp theo Năm / Tháng ──
    from collections import defaultdict
    agg = defaultdict(lambda: {
        "bt_san_luong": 0, "cd_san_luong": 0, "td_san_luong": 0,
        "bt_thanh_tien": 0, "cd_thanh_tien": 0, "td_thanh_tien": 0,
        "tien_chua_thue": 0, "tien_thue": 0, "tong_thanh_tien": 0,
    })
    for e in cleaned:
        key = (int(e["nam"]), int(e["thang"]))
        agg[key]["bt_san_luong"] += e["bt_san_luong"]
        agg[key]["cd_san_luong"] += e["cd_san_luong"]
        agg[key]["td_san_luong"] += e["td_san_luong"]
        agg[key]["bt_thanh_tien"] += e["bt_thanh_tien"]
        agg[key]["cd_thanh_tien"] += e["cd_thanh_tien"]
        agg[key]["td_thanh_tien"] += e["td_thanh_tien"]
        agg[key]["tien_chua_thue"] += e["tien_chua_thue"]
        agg[key]["tien_thue"] += e["tien_thue"]
        agg[key]["tong_thanh_tien"] += e["tong_thanh_tien"]

    # Nhóm theo năm
    years = sorted(set(k[0] for k in agg.keys()))

    # ── Styles ──
    fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_sub = Font(name="Calibri", size=10, bold=True, italic=True, color="FFFFFF")
    fill_total = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    font_total = Font(name="Calibri", size=11, bold=True)
    font_data = Font(name="Calibri", size=11)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    thick_bottom = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="medium", color="000000"),
    )

    wb = Workbook()
    first_sheet = True

    # ── Tạo sheet tổng hợp cho mỗi năm ──
    for year in years:
        if first_sheet:
            ws = wb.active
            ws.title = str(year)
            first_sheet = False
        else:
            ws = wb.create_sheet(title=str(year))

        # Header row 1
        ws.cell(row=1, column=1, value="Tháng")
        ws.merge_cells("A1:A2")
        ws.cell(row=1, column=2, value="Tiêu thụ điện (kWh)")
        ws.merge_cells("B1:E1")
        ws.cell(row=1, column=6, value="Chi phí tiền điện (VNĐ)")
        ws.merge_cells("F1:J1")

        # Header row 2
        sub_headers = ["", "Bình thường", "Cao điểm", "Thấp điểm", "TỔNG kWh",
                       "Bình thường", "Cao điểm", "Thấp điểm", "Thuế GTGT", "TỔNG THANH TOÁN"]
        for c, h in enumerate(sub_headers, 1):
            if h:
                ws.cell(row=2, column=c, value=h)

        # Format headers
        for r in range(1, 3):
            for c in range(1, 11):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill_header
                cell.font = font_header if r == 1 else font_sub
                cell.alignment = align_center
                cell.border = thin_border

        # Data rows (12 tháng)
        for month in range(1, 13):
            r = 2 + month
            key = (year, month)
            d = agg.get(key)

            ws.cell(row=r, column=1, value=month).alignment = align_center

            if d and (d["bt_san_luong"] + d["cd_san_luong"] + d["td_san_luong"]) > 0:
                bt_sl = d["bt_san_luong"]
                cd_sl = d["cd_san_luong"]
                td_sl = d["td_san_luong"]
                tong_sl = bt_sl + cd_sl + td_sl
                bt_tt = d["bt_thanh_tien"]
                cd_tt = d["cd_thanh_tien"]
                td_tt = d["td_thanh_tien"]
                tien_thue = d["tien_thue"]
                tong_tt = d["tong_thanh_tien"]

                ws.cell(row=r, column=2, value=bt_sl)
                ws.cell(row=r, column=3, value=cd_sl)
                ws.cell(row=r, column=4, value=td_sl)
                ws.cell(row=r, column=5, value=tong_sl)
                ws.cell(row=r, column=6, value=bt_tt)
                ws.cell(row=r, column=7, value=cd_tt)
                ws.cell(row=r, column=8, value=td_tt)
                ws.cell(row=r, column=9, value=tien_thue)
                ws.cell(row=r, column=10, value=tong_tt)

            for c in range(1, 11):
                cell = ws.cell(row=r, column=c)
                cell.font = font_data
                cell.border = thin_border
                if c >= 2:
                    cell.number_format = "#,##0"

        # Dòng TỔNG CỘNG
        tot_row = 15
        ws.cell(row=tot_row, column=1, value="TỔNG CỘNG").alignment = align_center
        for c in range(2, 11):
            col_letter = get_column_letter(c)
            ws.cell(row=tot_row, column=c, value=f"=SUM({col_letter}3:{col_letter}14)")
            ws.cell(row=tot_row, column=c).number_format = "#,##0"

        for c in range(1, 11):
            cell = ws.cell(row=tot_row, column=c)
            cell.font = font_total
            cell.fill = fill_total
            cell.border = thick_bottom

        # Auto-fit columns
        col_widths = [8, 14, 14, 14, 14, 16, 16, 16, 16, 18]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet Chi tiết từng Kỳ ──
    ws_detail = wb.create_sheet(title="Chi tiết từng Kỳ")
    detail_headers = ["Năm", "Tháng", "Kỳ", "Thuế (%)",
                      "ĐG BT", "SL BT (kWh)", "Thành tiền BT",
                      "ĐG CĐ", "SL CĐ (kWh)", "Thành tiền CĐ",
                      "ĐG TĐ", "SL TĐ (kWh)", "Thành tiền TĐ",
                      "Tổng kWh", "Tiền chưa thuế", "Thuế GTGT", "Tổng thanh toán", "Ghi chú"]
    ws_detail.append(detail_headers)
    for c_idx in range(1, len(detail_headers) + 1):
        cell = ws_detail.cell(row=1, column=c_idx)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border

    sorted_data = sorted(cleaned, key=lambda x: (x["nam"], x["thang"], x["ky"]))
    for row_idx, e in enumerate(sorted_data, 2):
        ws_detail.append([
            int(e["nam"]), int(e["thang"]), int(e["ky"]),
            e["thue_vat"],
            e["bt_don_gia"], e["bt_san_luong"], e["bt_thanh_tien"],
            e["cd_don_gia"], e["cd_san_luong"], e["cd_thanh_tien"],
e["td_don_gia"], e["td_san_luong"], e["td_thanh_tien"],
            e["tong_san_luong"], e["tien_chua_thue"], e["tien_thue"], e["tong_thanh_tien"],
            e["ghi_chu"]
        ])
        for c in range(1, len(detail_headers) + 1):
            cell = ws_detail.cell(row=row_idx, column=c)
            cell.font = font_data
            cell.border = thin_border
            if c <= 4:
                cell.alignment = align_center
            elif c <= 17:
                cell.number_format = "#,##0"

    for col in ws_detail.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_detail.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # ── Trả file ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=output_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def parse_electricity_text_content(raw_text: str) -> list:
    """
    Thuật toán phân tích và bóc tách dữ liệu điện 3 giá chuyên sâu từ file văn bản (.txt) hoặc kết quả OCR.
    
    Hỗ trợ:
    1. Bóc tách nhiều trạm biến áp / công tơ (PB...) trên cùng 1 file.
    2. Nhận diện chính xác Kỳ / Tháng / Năm từ tên file (# FILE: ...), tiêu đề hóa đơn hoặc khoảng thời gian ghi chỉ số.
    3. Nhận diện Thuế suất VAT (8%, 10%, 5%, 0%).
    4. Tự động nhận diện và gán đúng 3 mức Đơn giá & Sản lượng (Bình thường, Cao điểm, Thấp điểm) 
       kể cả khi bảng có cấu trúc hàng ngang, cột dọc hoặc OCR ngắt dòng.
    5. Fallback nhận diện cho dạng dữ liệu dán trực tiếp 3 dòng số từ bảng kê.
    """
    if not raw_text or not raw_text.strip():
        return []

    def clean_num(s):
        if not s:
            return 0.0
        s = str(s).strip().replace(" ", "").replace("\xa0", "").replace("\u200b", "").replace("\ufeff", "")
        # Nếu có cả dấu chấm và dấu phẩy (VD: 1.508.795,647 hoặc 1,508,795.647)
        if "." in s and "," in s:
            if s.rfind(",") > s.rfind("."):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        elif "," in s:
            parts = s.split(",")
            if len(parts) == 2 and len(parts[1]) in [1, 2, 3]:
                s = s.replace(",", ".")
            else:
                s = s.replace(",", "")
        elif "." in s:
            parts = s.split(".")
            if len(parts) > 2:
                s = s.replace(".", "")
            elif len(parts) == 2 and len(parts[1]) == 3:
                s = s.replace(".", "")
        try:
            return float(s)
        except ValueError:
            return 0.0

    # 1. Tách theo từng tệp (# FILE: ...) hoặc trang
    file_pattern = re.compile(r'(?:^|\n)#{5,}\s*\n#\s*FILE:\s*([^\n#]+)\s*\n#{5,}', re.IGNORECASE)
    splits = list(file_pattern.finditer(raw_text))

    blocks = []
    if splits:
        for i in range(len(splits)):
            fname = splits[i].group(1).strip()
            start = splits[i].end()
            end = splits[i+1].start() if i + 1 < len(splits) else len(raw_text)
            chunk = raw_text[start:end].strip()
            if chunk:
                blocks.append((fname, chunk))
    else:
        # Tách theo trang hoặc giữ nguyên
        page_chunks = re.split(r'={3,}\s*\[Trang[^\n]+\]\s*={3,}', raw_text, flags=re.IGNORECASE)
        if len(page_chunks) > 1:
            for p in page_chunks:
                p_str = p.strip()
                if p_str:
                    blocks.append(('', p_str))
        else:
            blocks = [('', raw_text.strip())]

    extracted_rows = []

    for filename, block_text in blocks:
        # Nếu trong 1 file/trang có nhiều trạm (PB...) hoặc nhiều bảng thanh toán, tách nhỏ
        station_splits = re.split(r'(?=(?:M[aã]\s*kh[aá]ch\s*h[aà]ng|M[aã]\s*tr[aạ]m|M[aã]\s*[ĐD]L|M[aã]\s*C[OÔ]NG\s*T[OƠ])[:\s]*(?:PB|PE|PA|PK|PD)\d+)', block_text, flags=re.IGNORECASE)
        if len(station_splits) <= 1:
            station_splits = [block_text]

        for st_text in station_splits:
            text = st_text.strip()
            if len(text) < 15:
                continue

            # --- A. Xác định Kỳ / Tháng / Năm ---
            ky, thang, nam = 1, 1, 2025
            found_period = False

            # 1. Từ câu: Kỳ hóa đơn: Kỳ 3 - 3/2026 hoặc Ky 1 - 01/2024
            m_ky = re.search(r'(?:K[yỳ]|Ky)\s*(?:h[oó]a\s*[đd][oơ]n|hoa\s*don)?[:\s]*(?:K[yỳ]|Ky)?\s*(\d+)\s*[-_/]\s*(\d{1,2})[/_-](\d{4})', text, re.IGNORECASE)
            if m_ky:
                ky = int(m_ky.group(1))
                thang = int(m_ky.group(2))
                nam = int(m_ky.group(3))
                found_period = True

            # 2. Từ tên file (VD: 1-1-24.pdf, Ky_1_Thang_01_2024.pdf)
            if not found_period and filename:
                m_fn = re.search(r'(?:^|[^\d])(\d{1,2})[-_](\d{1,2})[-_](\d{2,4})(?:[^\d]|$)', filename)
                if m_fn:
                    ky = int(m_fn.group(1))
                    thang = int(m_fn.group(2))
                    raw_y = int(m_fn.group(3))
                    nam = 2000 + raw_y if raw_y < 100 else raw_y
                    found_period = True


            # 3. Từ khoảng ngày: Từ ngày 01/01/2024 đến 10/01/2024
            if not found_period:
                m_range = re.search(r'T[uừ]\s*ng[aà]y\s*(\d{1,2})[/_-](\d{1,2})[/_-](\d{4})\s*[đd][eế]n\s*ng[aà]y\s*(\d{1,2})[/_-](\d{1,2})[/_-](\d{4})', text, re.IGNORECASE)
                if m_range:
                    start_day = int(m_range.group(1))
                    thang = int(m_range.group(2))
                    nam = int(m_range.group(3))
                    if start_day <= 10:
                        ky = 1
                    elif start_day <= 20:
                        ky = 2
                    else:
                        ky = 3
                    found_period = True

            # 4. Từ định dạng tiếng Anh: Period III 2026 hoặc March-Period III 2026
            if not found_period:
                m_p = re.search(r'Period\s+(I{1,3}|IV|\d+)\s+(\d{4})', text, re.IGNORECASE)
                if m_p:
                    p_map = {'I': 1, 'II': 2, 'III': 3, 'IV': 4}
                    ky = p_map.get(m_p.group(1).upper(), 1)
                    nam = int(m_p.group(2))
                    m_m = re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[-_\s]+Period', text, re.IGNORECASE)
                    if m_m:
                        m_dict = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6, 'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}
                        thang = m_dict.get(m_m.group(1).lower()[:3], 1)
                    found_period = True

            # 5. Từ tháng / năm thông thường
            if not found_period:
                m_th = re.search(r'th[aá]ng\s*(\d{1,2})\s*n[aă]m\s*(\d{4})', text, re.IGNORECASE)
                if m_th:
                    thang = int(m_th.group(1))
                    nam = int(m_th.group(2))

            # --- B. Xác định Thuế VAT (%) ---
            thue_vat = 0.0
            m_tax = re.search(r'(?:Thu[eếo]\s*su[aáâăắặẳẵấầẩẫậ]*t\s*GTGT|Thue\s*suat\s*GTGT|Thu[eế]\s*GTGT|VAT)[\s:\n]*(\d+)\s*%', text, re.IGNORECASE)
            if m_tax:
                thue_vat = float(m_tax.group(1))
            elif '8%' in text or ' 8 %' in text or 'E%' in text:
                thue_vat = 8.0
            elif '10%' in text or ' 10 %' in text:
                thue_vat = 10.0
            elif '5%' in text or ' 5 %' in text:
                thue_vat = 5.0

            # --- C. Tìm Mã trạm / Mã Khách hàng ---
            m_kh = re.search(r'((?:PB|PE|PA|PK|PD)\d{8,12})', text)
            kh_code = m_kh.group(1) if m_kh else ''

            # --- D. Trích xuất 3 khung giá (Bình thường, Cao điểm, Thấp điểm) ---
            def extract_tier_smart(regex_tier, content):
                matches = list(re.finditer(regex_tier, content, re.IGNORECASE))
                best_price, best_vol = 0.0, 0.0
                for m in matches:
                    start_pos = m.end()
                    chunk = content[start_pos:start_pos + 160]
                    # Ngắt nếu gặp từ khóa khung giờ tiếp theo hoặc phần tổng
                    chunk = re.split(r'(?:Khung\s*gi|Gi[oò]\s*cao|Gi[oò]\s*th|Gi[oò]\s*b|T[OỔ]NG\s*S[OỐ]|Thu[eếo]|100\s*%)', chunk, flags=re.IGNORECASE)[0]
                    nums = [clean_num(n) for n in re.findall(r'\d+[\d\.,]*', chunk) if len(n) > 0]
                    nums = [n for n in nums if n > 0]

                    if len(nums) >= 3:
                        # 3 số: [đơn_giá, sản_lượng, thành_tiền]
                        p, v, tot = nums[0], nums[1], nums[2]
                        if 800 <= p <= 6000:
                            return p, v
                        elif 800 <= nums[1] <= 6000:
                            return nums[1], nums[0]
                    elif len(nums) == 2:
                        n1, n2 = nums[0], nums[1]
                        if 800 <= n1 <= 6000:
                            return n1, n2
                        elif 800 <= n2 <= 6000:
                            return n2, n1
                        else:
                            return 0.0, max(n1, n2)
                    elif len(nums) == 1:
                        if nums[0] > 6000:
                            best_vol = nums[0]
                return best_price, best_vol

            # Ưu tiên lấy vùng bảng thanh toán (sau TỔNG SỐ TIỀN THANH TOÁN hoặc THÀNH TIỀN)
            sub_search = text
            m_sub = re.search(r'T[OỔ]NG\s*S[OỐ]\s*TI[EỀ]N\s*THANH\s*TO[AÁ]N|TH[AÀ]NH\s*TI[EỀ]N\s*\([đd][oồ]ng\)|TI[EỀ]N\s*[ĐD]I[EỆ]N\s*CHI\s*TI[EẾ]T', text, re.IGNORECASE)
            if m_sub:
                sub_search = text[m_sub.start():]

            re_bt = r'(?:b[iìíỉĩị]nh\s*th[uư][oơờ][^\s]*|binh\s*thuong|BT|Normal)'
            re_cd = r'(?:cao\s*[dđ][ií][^\s]*|cao\s*diem|CD|CĐ|Peak)'
            re_td = r'(?:th[aáâăắặẳẵấầẩẫậ]p\s*[dđ][ií][^\s]*|thap\s*diem|TD|TĐ|Off-peak)'

            bt_p, bt_v = extract_tier_smart(re_bt, sub_search)
            cd_p, cd_v = extract_tier_smart(re_cd, sub_search)
            td_p, td_v = extract_tier_smart(re_td, sub_search)

            # Fallback nếu không có trong sub_search thì tìm toàn văn
            if (bt_v + cd_v + td_v) == 0:
                bt_p, bt_v = extract_tier_smart(re_bt, text)
                cd_p, cd_v = extract_tier_smart(re_cd, text)
                td_p, td_v = extract_tier_smart(re_td, text)

            if (bt_v + cd_v + td_v) > 0:
                note = f"OCR {kh_code}" if kh_code else (filename if filename else "OCR txt")
                extracted_rows.append({
                    "nam": nam,
                    "thang": thang,
                    "ky": ky,
                    "thue_vat": thue_vat,
                    "bt_don_gia": bt_p or 1833.0,
                    "bt_san_luong": bt_v,
                    "cd_don_gia": cd_p or 3398.0,
                    "cd_san_luong": cd_v,
                    "td_don_gia": td_p or 1190.0,
                    "td_san_luong": td_v,
                    "ghi_chu": note
                })

    # --- Fallback: Nhận diện dạng dán trực tiếp 3 dòng số (1.833 51.559 ...) ---
    if len(extracted_rows) == 0:
        lines = [l.strip() for l in raw_text.split('\n') if l.strip()]
        for i in range(0, len(lines) - 2, 3):
            group_nums = []
            for j in range(3):
                nums = [clean_num(n) for n in re.findall(r'\d+[\d\.,]*', lines[i + j]) if clean_num(n) > 0]
                if len(nums) >= 2:
                    p, v = (nums[0], nums[1]) if 800 <= nums[0] <= 6000 else (nums[1], nums[0])
                    group_nums.append((p, v))
                elif len(nums) == 1:
                    group_nums.append((0.0, nums[0]))
            if len(group_nums) == 3 and (group_nums[0][1] + group_nums[1][1] + group_nums[2][1]) > 0:
                extracted_rows.append({
                    "nam": 2025,
                    "thang": 1,
                    "ky": 1,
                    "thue_vat": 0.0,
                    "bt_don_gia": group_nums[0][0] or 1833.0,
                    "bt_san_luong": group_nums[0][1],
                    "cd_don_gia": group_nums[1][0] or 3398.0,
                    "cd_san_luong": group_nums[1][1],
                    "td_don_gia": group_nums[2][0] or 1190.0,
                    "td_san_luong": group_nums[2][1],
                    "ghi_chu": "text paste"
                })

    extracted_rows.sort(key=lambda x: (x["nam"], x["thang"], x["ky"]))
    return extracted_rows


@excel_bp.route('/parse-invoices', methods=['POST'])
def parse_invoices():
    """
    Tự động đọc và bóc tách dữ liệu điện 3 giá từ file ZIP hoặc nhiều file PDF hóa đơn EVN (dạng vector/digital PDF).
    
    Returns:
        JSON response chứa mảng dataRows đã bóc tách.
    """
    try:
        import pdfplumber
    except ImportError:
        return jsonify({"error": "Thiếu thư viện pdfplumber trên server."}), 500

    uploaded_files = request.files.getlist("files")
    if not uploaded_files or (len(uploaded_files) == 1 and not uploaded_files[0].filename):
        single_file = request.files.get("file")
        if single_file and single_file.filename:
            uploaded_files = [single_file]

    if not uploaded_files or not uploaded_files[0].filename:
        return jsonify({"error": "Vui lòng chọn file ZIP hoặc PDF hóa đơn."}), 400

    def extract_pdfs_from_bytes(file_bytes: bytes, fname: str) -> list:
        pdf_list = []
        if fname.lower().endswith(".pdf"):
            pdf_list.append((fname, file_bytes))
        elif fname.lower().endswith(".zip"):
            try:
                with zipfile.ZipFile(io.BytesIO(file_bytes), "r") as zf:
                    pdf_names = [name for name in zf.namelist() if name.lower().endswith(".pdf") and not name.startswith("__MACOSX")]
                    ctiet_names = [n for n in pdf_names if "_ctiet" in n.lower() or "_ct." in n.lower()]
                    target_names = ctiet_names if len(ctiet_names) > 0 else pdf_names
                    for pname in target_names:
                        pdf_content = zf.read(pname)
                        pdf_list.append((os.path.basename(pname), pdf_content))
            except Exception:
                pass
        return pdf_list

    all_pdf_items = []
    for f in uploaded_files:
        f_bytes = f.read()
        extracted = extract_pdfs_from_bytes(f_bytes, f.filename)
        all_pdf_items.extend(extracted)

    if not all_pdf_items:
        return jsonify({"error": "Không tìm thấy file PDF hóa đơn nào trong tệp đã tải lên."}), 400

    parsed_rows = []
    failed_files = []

    for pdf_name, pdf_bytes in all_pdf_items:
        try:
            text = ""
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as doc:
                for page in doc.pages:
                    t = page.extract_text()
                    if t:
                        text += t + "\n"

            if not text.strip():
                raise ValueError(f"Không tìm thấy lớp văn bản (vector text) trong file {pdf_name}. Với file scan, vui lòng sử dụng Tool Desktop OCR.")

            # Chạy qua bộ phân tích chuyên sâu
            doc_text = f"########################################\n# FILE: {pdf_name}\n########################################\n\n{text}"
            rows = parse_electricity_text_content(doc_text)
            if not rows:
                raise ValueError(f"Không bóc tách được số liệu điện 3 giá từ nội dung file {pdf_name}.")
            parsed_rows.extend(rows)
        except Exception as e:
            failed_files.append({"file": pdf_name, "error": str(e)})

    parsed_rows.sort(key=lambda x: (x["nam"], x["thang"], x["ky"]))

    return jsonify({
        "success": True,
        "total_pdfs": len(all_pdf_items),
        "parsed_count": len(parsed_rows),
        "failed_count": len(failed_files),
        "failed_files": failed_files,
        "data": parsed_rows
    })


@excel_bp.route('/parse-text', methods=['POST'])
def parse_text_endpoint():
    """
    Phân tích văn bản thuần (plaintext từ OCR hoặc copy/paste) và bóc tách các dòng số liệu điện 3 giá.
    """
    payload = request.get_json(silent=True) or {}
    raw_text = payload.get('text', '')
    if not raw_text and 'text' in request.form:
        raw_text = request.form.get('text', '')

    if not raw_text or not raw_text.strip():
        if 'file' in request.files:
            raw_text = request.files['file'].read().decode('utf-8', errors='ignore')

    if not raw_text or not raw_text.strip():
        return jsonify({"error": "Văn bản rỗng hoặc không có dữ liệu để phân tích."}), 400

    extracted_rows = parse_electricity_text_content(raw_text)

    if not extracted_rows:
        return jsonify({"error": "Không nhận diện được cấu trúc số liệu điện 3 giá từ văn bản."}), 400

    return jsonify({
        "success": True,
        "count": len(extracted_rows),
        "data": extracted_rows
    })



